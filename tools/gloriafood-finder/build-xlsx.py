#!/usr/bin/env python3
"""Merge GloriaFood detection results with BuiltWith email/metadata into a clean XLSX."""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
LEADS = os.path.join(HERE, "leads")

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

# --- Load BuiltWith metadata (domain -> fields) ---------------------------
meta = {r["domain"].strip().lower(): r for r in read_csv(os.path.join(LEADS, "builtwith-extract.csv"))}

# --- Load detection results, merging the 3 runs (best outcome wins) -------
PRIORITY = {"detected": 3, "none": 2, "unreachable": 1}
def outcome_rank(row):
    if row["detected"] == "true":
        return 3
    return PRIORITY.get(row["confidence"], 0)

best = {}
for fname in ["results.csv", "retry-results.csv", "retry2-results.csv"]:
    p = os.path.join(LEADS, fname)
    if not os.path.exists(p):
        continue
    for row in read_csv(p):
        d = row["domain"].strip().lower()
        if d not in best or outcome_rank(row) > outcome_rank(best[d]):
            best[d] = row

# --- Build merged records in original list order --------------------------
ORDER = [r["domain"].strip().lower() for r in read_csv(os.path.join(LEADS, "builtwith-extract.csv"))]

def status_label(r):
    if r["detected"] == "true":
        return "GloriaFood — LIVE"
    if r["confidence"] == "unreachable":
        return "Unreachable — verify manually"
    return "No GloriaFood found"

records = []
for d in ORDER:
    m = meta.get(d, {})
    r = best.get(d, {"detected": "false", "confidence": "unreachable", "methods": "",
                     "cuid": "", "ruid": "", "matchedUrl": "", "httpStatus": "", "error": ""})
    status = status_label(r)
    note = ""
    if status.startswith("Unreachable"):
        note = f"Could not reach site ({r.get('error','')}). Check the domain manually."
    elif status.startswith("No GloriaFood"):
        note = "No GloriaFood widget in page source — likely removed it or switched provider."
    records.append({
        "domain": d,
        "company": m.get("company", ""),
        "email": m.get("email", ""),
        "phone": m.get("phone", ""),
        "city": m.get("city", ""),
        "postcode": m.get("postcode", ""),
        "country": m.get("country", ""),
        "cms": m.get("cms", ""),
        "status": status,
        "confidence": r.get("confidence", ""),
        "cuid": r.get("cuid", ""),
        "ruid": r.get("ruid", ""),
        "methods": r.get("methods", ""),
        "url": r.get("matchedUrl", "") or (f"https://{d}/" if not status.startswith("Unreachable") else ""),
    })

# Sort: LIVE first, then (has email first), then No GloriaFood, then Unreachable; alpha within.
status_order = {"GloriaFood — LIVE": 0, "No GloriaFood found": 1, "Unreachable — verify manually": 2}
records.sort(key=lambda x: (status_order[x["status"]], 0 if x["email"] else 1, x["domain"]))

# --- Write workbook --------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Leads"

HEADERS = [
    ("Restaurant / Domain", "domain", 30),
    ("Company", "company", 30),
    ("Email", "email", 34),
    ("Phone", "phone", 26),
    ("City", "city", 18),
    ("Postcode", "postcode", 12),
    ("Country", "country", 9),
    ("Website / CMS", "cms", 22),
    ("GloriaFood Status", "status", 26),
    ("GloriaFood Customer ID (cuid)", "cuid", 30),
    ("Restaurant ID (ruid)", "ruid", 30),
    ("Detection Evidence", "methods", 48),
    ("Site URL", "url", 38),
]

# Styles
brand = "1F2937"
header_fill = PatternFill("solid", fgColor=brand)
header_font = Font(bold=True, color="FFFFFF", size=11)
live_fill = PatternFill("solid", fgColor="E7F5EC")     # light green
none_fill = PatternFill("solid", fgColor="FDECEC")     # light red
unr_fill  = PatternFill("solid", fgColor="FEF6E0")     # light amber
email_font = Font(color="1155CC", underline="single")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
t = ws.cell(1, 1, "GloriaFood Migration Leads — tested for live widget  ·  Posso  ·  generated from BuiltWith export")
t.font = Font(bold=True, size=13, color=brand)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 26

# Header row
hr = 2
for c, (label, _, width) in enumerate(HEADERS, 1):
    cell = ws.cell(hr, c, label)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = width
ws.row_dimensions[hr].height = 30

# Data rows
for i, rec in enumerate(records):
    r = hr + 1 + i
    fill = live_fill if rec["status"].startswith("GloriaFood") else (unr_fill if rec["status"].startswith("Unreachable") else none_fill)
    for c, (_, key, _) in enumerate(HEADERS, 1):
        val = rec[key]
        cell = ws.cell(r, c, val)
        cell.fill = fill
        cell.border = border
        cell.alignment = wrap if key in ("methods", "email", "phone") else top
        if key == "email" and val:
            first = val.split(";")[0].strip()
            cell.hyperlink = f"mailto:{first}"
            cell.font = email_font
        if key == "url" and val:
            cell.hyperlink = val
            cell.font = email_font

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(HEADERS))}{hr+len(records)}"

# --- Summary sheet ---------------------------------------------------------
s = wb.create_sheet("Summary")
n_total = len(records)
n_live = sum(1 for r in records if r["status"].startswith("GloriaFood"))
n_none = sum(1 for r in records if r["status"].startswith("No GloriaFood"))
n_unr = sum(1 for r in records if r["status"].startswith("Unreachable"))
n_email = sum(1 for r in records if r["email"])
n_live_email = sum(1 for r in records if r["status"].startswith("GloriaFood") and r["email"])
n_ids = sum(1 for r in records if r["cuid"] or r["ruid"])

rows = [
    ("GloriaFood Migration Leads — Summary", ""),
    ("", ""),
    ("Domains tested", n_total),
    ("Confirmed LIVE GloriaFood widget", n_live),
    ("  ...of those, with an email", n_live_email),
    ("  ...of those, with extracted GloriaFood IDs", n_ids),
    ("No GloriaFood found (reachable)", n_none),
    ("Unreachable — verify manually", n_unr),
    ("", ""),
    ("Total emails extracted", n_email),
    ("", ""),
    ("Tested with", "tools/gloriafood-finder/find-gloriafood.mjs (ewm2.js + data-glf-* fingerprint)"),
    ("GloriaFood shutdown date", "30 April 2027 (Oracle)"),
    ("", ""),
    ("COMPLIANCE", "BuiltWith terms forbid using this list for BULK email. UK PECR: cold email OK to"),
    ("", "limited companies (Ltd/LLP) only; sole traders need phone/LinkedIn/post/visit."),
    ("", "Screen all phone numbers against TPS + CTPS first. See LEGAL-OUTREACH.md."),
]
for i, (k, v) in enumerate(rows, 1):
    a = s.cell(i, 1, k); b = s.cell(i, 2, v)
    if i == 1:
        a.font = Font(bold=True, size=14, color=brand)
    elif k in ("COMPLIANCE",):
        a.font = Font(bold=True, color="B45309")
    elif k and not k.startswith(" ") and v != "":
        a.font = Font(bold=True)
s.column_dimensions["A"].width = 42
s.column_dimensions["B"].width = 86

out = os.path.join(LEADS, "GloriaFood-Leads-Posso.xlsx")
wb.save(out)
print(f"Wrote {out}")
print(f"  {n_total} domains | {n_live} LIVE GloriaFood ({n_live_email} with email) | {n_none} none | {n_unr} unreachable | {n_email} emails total")
